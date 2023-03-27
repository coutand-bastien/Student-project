__name__      = "parser.py"
__author__    = "MARCHAND Robin"
__date__      = "05.10.22"


# ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~
#                                              ~~
#                    Imports                   ~~
#                                              ~~
# ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~

import os
import datetime as dt
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, PatternFill, Alignment

from api.parser.ICalendarGetter import ICalendarGetter
from api.parser.ICalendarParser import ICalendarParser


# ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~
#                                              ~~
#                    Corps                     ~~
#                                              ~~
# ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~


def create_excel_file(start_date, end_date, seuil, list_unique_rooms_without_multi_rooms, cal):
    """_summary_ : Fonction qui permet de créer un fichier excel avec les informations des événements

    Args:
        start_date (date): date de début
        end_date (date): date de fin
        seuil (int): seuil qui génère des couleurs adéquates à la saturation des salles

    Returns:
        _type_: fichier excel
    """

    try:
        seuil = int(seuil)
        if seuil < 0:
            seuil = 0
    except ValueError:
        seuil = 0


    seuil_alerte = 100 - seuil
    print("[+] seuil_alerte = " + str(seuil_alerte) + "%")

    seuil_alerte = seuil_alerte / 100

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    wb = Workbook()
    ws = wb.active
    max_row = (end_date - start_date).days + 1
    max_event = len(list_unique_rooms_without_multi_rooms)

    # Create header
    ws['A1'] = "Date"
    ws['B1'] = "8:00 - 9:30"
    ws['C1'] = "9:45 - 11:15"
    ws['D1'] = "11:30 - 13:00"
    ws['E1'] = "13h00 - 14h30"
    ws['F1'] = "14h45 - 16h15"
    ws['G1'] = "16h30 - 18h00"
    ws['H1'] = "18h15 - 19h45"


    for day in range(1, (end_date - start_date).days + 1):
        ws[f"B{day+1}"] = len([event for event in cal if event['DTSTART'].date() == start_date + dt.timedelta(days=day) and event['DTSTART'].time() >= dt.time(8, 0, 0) and event['DTEND'].time() <= dt.time(9, 30, 0)])
        ws[f"B{day+1}"] = ws[f"B{day+1}"].value + len([event for event in cal if event['DTSTART'].date() == start_date + dt.timedelta(days=day) and event['DTSTART'].time() >= dt.time(8, 30, 0) and event['DTEND'].time() <= dt.time(10, 0, 0)])
    
        ws[f"C{day+1}"] = len([event for event in cal if event['DTSTART'].date() == start_date + dt.timedelta(days=day) and event['DTSTART'].time() >= dt.time(9, 45, 0) and event['DTEND'].time() <= dt.time(11, 15, 0)])
        ws[f"D{day+1}"] = len([event for event in cal if event['DTSTART'].date() == start_date + dt.timedelta(days=day) and event['DTSTART'].time() >= dt.time(11, 30, 0) and event['DTEND'].time() <= dt.time(13, 00, 0)])
        ws[f"E{day+1}"] = len([event for event in cal if event['DTSTART'].date() == start_date + dt.timedelta(days=day) and event['DTSTART'].time() >= dt.time(13, 00, 0) and event['DTEND'].time() <= dt.time(14, 30, 0)])
        ws[f"F{day+1}"] = len([event for event in cal if event['DTSTART'].date() == start_date + dt.timedelta(days=day) and event['DTSTART'].time() >= dt.time(14, 45, 0) and event['DTEND'].time() <= dt.time(16, 15, 0)])
        ws[f"G{day+1}"] = len([event for event in cal if event['DTSTART'].date() == start_date + dt.timedelta(days=day) and event['DTSTART'].time() >= dt.time(16, 30, 0) and event['DTEND'].time() <= dt.time(18, 00, 0)])
        ws[f"H{day+1}"] = len([event for event in cal if event['DTSTART'].date() == start_date + dt.timedelta(days=day) and event['DTSTART'].time() >= dt.time(18, 15, 0) and event['DTEND'].time() <= dt.time(19, 45, 0)])
        
        ws[f"A{day+1}"] = start_date + dt.timedelta(days=day)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20

        ws[f"A{day+1}"].number_format = 'ddd dd/mm/yyyy'

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=1):
        for cell in row:
            cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=8):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=8):
        if sum([cell.value for cell in row]) == 0:
            for cell in row:
                cell.fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
                cell.value = None
                    
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=8):
        for cell in row:
            try : 
                if cell.value > seuil_alerte * max_event:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            except TypeError:
                pass

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=8):
        for cell in row:
            try : 
                if cell.value > seuil_alerte * max_event * 0.6 and cell.value <= seuil_alerte * max_event:
                    cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            except TypeError:
                pass

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=8):
        for cell in row:
            try : 
                if cell.value > seuil_alerte * max_event * 0.2 and cell.value <= seuil_alerte * max_event * 0.6:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            except TypeError:
                pass

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=8):
        for cell in row:
            try : 
                if cell.value <= seuil_alerte * max_event * 0.2:
                    cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
            except TypeError:
                pass


        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')


        ws['J10'] = "Seuil d'alerte :"
        ws['J10'].font = Font(bold=True)
        ws.column_dimensions['J'].width = 30
        ws['J10'].alignment = Alignment(horizontal='center', vertical='center')

        ws['J11'] = f"{(seuil)} %"


        ws['J11'].alignment = Alignment(horizontal='center', vertical='center')
        #Shrink to fit
        ws.column_dimensions['J'].width = 20


        ws['J16'] = "Date de début :"
        # In bold
        ws['J16'].font = Font(bold=True)
        ws.column_dimensions['J'].width = 30

        ws['J16'].alignment = Alignment(horizontal='center', vertical='center')

        ws['J17'] = start_date + timedelta(days=1)
        ws.column_dimensions['J'].width = 20

        ws['J17'].alignment = Alignment(horizontal='center', vertical='center')

        ws['J19'] = "Date de fin :"
        ws['J19'].font = Font(bold=True)
        ws.column_dimensions['J'].width = 30

        ws['J19'].alignment = Alignment(horizontal='center', vertical='center')

        ws['J20'] = end_date
        ws.column_dimensions['J'].width = 20

        ws['J20'].alignment = Alignment(horizontal='center', vertical='center')

        # Save workbook
    wb.save("api/parser/excels/output.xlsx")

    print("==> Done (saved in excels/output.xlsx)")
    print("Creation of an excel which contains : " + str((end_date - start_date).days) + " days")

    print("=> Start date : " + str(start_date + timedelta(days=1)))
    print("=> End date : " + str(end_date))
    print("=> Seuil d'alerte : " + str(seuil) + " %")
    return wb


# ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~
#                                              ~~
#                    MAIN                      ~~
#                                              ~~
# ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~

def parser():
    """_summary_ : Main function"""
    
    url = 'https://planning.univ-ubs.fr/jsp/custom/modules/plannings/anonymous_cal.jsp?data=d3be1fff3397511bf2829d59289190e54592592d3b282f749c9a606b710f264fdc5c094f7d1a811b903031bde802c7f597e5f20f622768ff000996ffaaa109f62d0bb596bf80452fd9b95e257006d2f8166c54e36382c1aa3eb0ff5cb8980cdb,1'
    ensibs = ICalendarGetter(url, 'api/parser/ics/ensibs.ics')

    ensibs_calendar = ICalendarParser(ensibs.get_calendar())
    cal = ensibs_calendar.parse()

    """
    Print the calendar with the events, sorted by date, and formatted to be readable
    for event in sorted(cal, key=lambda x: x['DTSTART']):
    print(f"{event['DTSTART'].strftime('%d/%m/%Y %H:%M')} - {event['DTEND'].strftime('%H:%M')} : {event['LOCATION']}")
    """

    print("------------------ DATA ------------------")

    list_rooms = [event['LOCATION'] for event in cal]
    list_unique_rooms = set(list_rooms)
    list_multi_rooms = [room for room in list_rooms if "," in room]
    list_unique_multi_rooms = set(list_multi_rooms)
    list_single_rooms = [room for room in list_rooms if "," not in room]
    list_unique_single_rooms = set(list_single_rooms)
    list_unique_rooms_without_multi_rooms = list_unique_rooms - list_unique_multi_rooms
    list_unique_rooms_with_multi_rooms = list_unique_rooms - list_unique_rooms_without_multi_rooms


    with open('api/parser/list_rooms.txt', 'w') as f:
        for room in sorted(list_unique_rooms_without_multi_rooms):
            f.write(room + "\n")
    print("==> Done (saved in list_rooms.txt)")



    print("[+] Number of multi-rooms : " + str(len(list_unique_multi_rooms)))
    print("[+] Number of single rooms : " + str(len(list_unique_rooms_without_multi_rooms)))
    print("[+] Number of rooms with multi-rooms : " + str(len(list_unique_rooms_with_multi_rooms)))


    with open('api/parser/list_events.txt', 'w') as f:
        for event in sorted(cal, key=lambda x: x['DTSTART']):
            if event['DTSTART'].date() == datetime(2022, 9, 6).date():
                f.write(f"{event['DTSTART'].strftime('%d/%m/%Y %H:%M')} - {event['DTEND'].strftime('%H:%M')} : {event['LOCATION']}" + "\n") 
    print("==> Done (saved in list_events.txt)")

    with open('api/parser/list_events_before_8.txt', 'w') as f:
        for event in sorted(cal, key=lambda x: x['DTSTART']):

            if event['DTSTART'].date() == datetime(2022, 9, 6).date() and event['DTSTART'].time() < datetime(2022, 9, 6, 8, 0).time():
                f.write(f"{event['DTSTART'].strftime('%d/%m/%Y %H:%M')} - {event['DTEND'].strftime('%H:%M')} : {event['LOCATION']}" + "\n")
    print("==> Done (saved in list_events_before_8.txt)")
    

    list_days = [event['DTSTART'].date() for event in cal]
    list_unique_days = set(list_days)
    list_days_with_number_of_events = [(day, list_days.count(day)) for day in list_unique_days]
    list_days_with_number_of_events = sorted(list_days_with_number_of_events, key=lambda x: x[1], reverse=True)


    # ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~
    #                                              ~~
    #                Appel EXCEL                   ~~
    #                                              ~~
    # ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~

    with open('api/parser/config/configExcel.txt', 'r') as f:
        datedeb = f.readline()
        datefin = f.readline()
        seuil = f.readline()

    datedeb = datedeb[:-1]
    datefin = datefin[:-1]
    seuil = seuil[:-1]

    datedeb = str(datedeb)
    datefin = str(datefin)
    seuil = int(seuil)

    date1 = datetime.strptime(datedeb, "%Y, %m, %d")
    date2 = datetime.strptime(datefin, "%Y, %m, %d")
    date3 = seuil
    date1 = date1.date()
    date2 = date2.date()

    create_excel_file(date1, date2, date3, list_unique_rooms_without_multi_rooms, cal)

'''   
    os.system("libreoffice --calc api/parser/excels/output.xlsx")


    # ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~
    #                                              ~~
    #                Appel PIE                     ~~
    #                                              ~~
    # ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~



    with open('config/configPie.txt', 'r') as f:
        dateASK = f.readline()
        sallePIE = f.readline()

    dateASK = dateASK[:-1]
    sallePIE = sallePIE[:-1]

    dateASK = str(dateASK)
    sallePIE = str(sallePIE)

    date1PIE = datetime.strptime(dateASK, "%Y, %m, %d")
    date1PIE = date1PIE.date()

    #create_pie_chart(date1PIE, sallePIE)

    # ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~
    #                                              ~~
    #              Appel OccupiedOrNot             ~~
    #                                              ~~
    # ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~ ~~

    with open('config/configOcc.txt', 'r') as f:
        heure = f.readline()
        salle = f.readline()
    
    heure = heure[:-1]
    salle = salle[:-1]

    heure = str(heure)
    salle = str(salle)

    dateOCCUPATION = datetime.strptime(heure, "%Y, %m, %d, %H, %M")
    print(dateOCCUPATION)

    #is_room_occupied(salle, dateOCCUPATION)

def create_pie_chart(day, room):
    """_summary_ : Create a pie chart with the number of events per category

    Args:
        day (str): Date of the day
        room (str): Room name

    Returns:
        _type_ : _description_
    """
    if room not in [event['LOCATION'] for event in cal]:
        return print("Error : The room you entered doesn't match with the room in the calendar")

    import matplotlib.pyplot as plt
    fig = plt.figure(figsize=(10, 7))

    all_possible_creneau_occupied = len([event for event in cal if event['DTSTART'].date() == day and event['LOCATION'] == room])
    plt.pie([all_possible_creneau_occupied, 7 - all_possible_creneau_occupied], labels=[f"{all_possible_creneau_occupied} x fois occupée", f"{7 - all_possible_creneau_occupied} x fois libre"], colors=["red", "green"], autopct='%1.1f%%')
    plt.title(f"Occupation rate of the room {room} for the {day}")
    for event in cal:
        if event['DTSTART'].date() == day and event['LOCATION'] == room:
            print(f"/!\ Occupied from {event['DTSTART'].time()} to {event['DTEND'].time()}")
    # Save the pie chart
    plt.show()

    

def is_room_occupied(roomSearch, hour):
    from matplotlib import pyplot as plt
    if room not in list_unique_rooms_without_multi_rooms:
        return print("Error : The room you entered doesn't match with the room in the calendar")
    if len([event for event in cal if event['DTSTART'].date() == hour.date() and event['DTSTART'].time() <= hour.time() and event['DTEND'].time() >= hour.time() and event['LOCATION'] == roomSearch]) > 0:
        plt.text(0.5, 0.5, "Occupied", fontsize=50, color='red', horizontalalignment='center', verticalalignment='center')
        plt.text(0.5, 0.4, roomSearch, fontsize=20, color='black', horizontalalignment='center', verticalalignment='center')
        plt.text(0.5, 0.3, hour.strftime("%H:%M"), fontsize=20, color='black', horizontalalignment='center', verticalalignment='center')
        plt.text(0.5, 0.2, hour.strftime("%d/%m/%Y"), fontsize=20, color='black', horizontalalignment='center', verticalalignment='center')
        plt.show()
        return print(f"The room {roomSearch} is occupied at {hour}")
    else:
        plt.text(0.5, 0.5, "FREE", fontsize=50, color='green', horizontalalignment='center', verticalalignment='center')
        plt.text(0.5, 0.4, roomSearch, fontsize=20, color='black', horizontalalignment='center', verticalalignment='center')
        plt.text(0.5, 0.3, hour.strftime("%H:%M"), fontsize=20, color='black', horizontalalignment='center', verticalalignment='center')
        plt.text(0.5, 0.2, hour.strftime("%d/%m/%Y"), fontsize=20, color='black', horizontalalignment='center', verticalalignment='center')
        plt.show()
        return print(f"The room {roomSearch} is free at {hour}")
'''